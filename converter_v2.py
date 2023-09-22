import datetime
from enum import Enum
import json
from math import floor
import random
from openpyxl import Workbook, load_workbook

wb = Workbook()

# TODO Hacer un parser para cambiar la forma en la que los turnos son descritos a la forma en la que yo lo uso.


# Obtiene la semana de la tabla
def get_week(year, week_index, table_name: str) -> dict:
    days = []

    table = wb[table_name]

    # salto en filas para grupos de 5 semanas
    initial_row = 4 + 8 * floor(week_index / 5)

    # Número de la semana saltando entre cada 5 de estas
    week_number = week_index % 5

    # Posición de inicio de la semana especificada
    week_start_index = 4 + week_number * 5

    # Fecha de inicio de semana
    weel_raw_start_date = str(table.cell(
        initial_row - 2, week_start_index + 1).value)
    week_start_date = (parse_week_date(year, weel_raw_start_date)
                       if weel_raw_start_date != "None" else "")

    # Número de la semana
    raw_week_number = str(table.cell(initial_row - 2, week_start_index).value)
    week_number = int(raw_week_number if raw_week_number != "None" else -1)

    day_date = week_start_date

    # Por cada día de la semana
    for column_index in range(week_start_index, week_start_index + 5):

        events = []

        # Por cada evento del dia
        turn = 1
        for row_index in range(initial_row, initial_row + 6):

            event = (
                str(table.cell(row_index, column_index).value)
                if table.cell(row_index, column_index).value != None
                else ""
            )

            parsed_event = parse_event(turn, event)

            turn += 1

            if (event) == "":
                continue

            events.append(parsed_event)

        day = {
            "date": day_date if day_date != "PARSE_ERROR" else day_date,
            "events": events
        }

        days.append(day)

        if day_date != "":
            day_date_converted = datetime.datetime.strptime(
                day_date, "%Y-%m-%d") + datetime.timedelta(days=1)

            day_date = day_date_converted.strftime("%Y-%m-%d")

    week = {
        "number": week_number,
        "startDate": week_start_date,
        "days": days,
    }

    return week


month_conv = {
    "ene": "1",
    "feb": "2",
    "mar": "3",
    "abr": "4",
    "may": "5",
    "jun": "6",
    "jul": "7",
    "ago": "8",
    "sep": "9",
    "oct": "10",
    "nov": "11",
    "dic": "12",
}

last_month = ""
actual_year = 0


def parse_week_date(year: int, input: str):

    global actual_year
    global last_month

    if input == "":
        return input

    # Para 20 al 26/feb
    splitted_i = input.strip().split(" ")
    splitted_s = splitted_i[0].split("/")
    splitted_f = splitted_i[2].split("/")

    if len(splitted_s) < 2:
        splitted_s.append(splitted_f[1])

    s_f = [
        splitted_s[0].zfill(2),
        month_conv[splitted_s[1]],
        splitted_f[0].zfill(2),
        month_conv[splitted_f[1]]
    ]

    if last_month:
        if last_month == "Dec" and s_f[1] == "Jan":
            actual_year = year + 1

    if s_f[2] == "Dec" and s_f[4] == "Jan":
        actual_year = year + 1

    last_month = s_f[1]

    s_f.insert(0, str(actual_year))
    final = f"{s_f[0]}-{s_f[2].zfill(2)}-{s_f[1]}"

    return final


def get_weeks(year: int, table_name):
    is_empty = False
    current_index = 0

    weeks = []

    while is_empty == False:
        current_week = get_week(year, current_index, table_name)

        if current_week["startDate"] == "":
            break

        weeks.append(current_week)

        current_index += 1

    return weeks


def parse_title(input: str):
    title = input.strip().split(" ")

    index = 0
    while index < len(title):
        if title[index] == "" and title[index + 1] == "":
            del title[index]
            index -= 1

        index += 1

    result = ""
    for index in range(len(title)):
        if index > 0:
            result += " "
        result += parse_name(title[index])

    result_dict = result.split("  ")

    return result_dict


def parse_name(input: str):
    if input.count(":"):
        input = input.split(":")[1].strip()

    return input


subject_names = {
    "M": "Matemática II",
    "P": "Programación II",
    "D": "Matemática Discreta II",
    "U": "Lenguaje Unificado de Modelado (UML)",
    "H": "Historia de Cuba",
    "N": "Defensa Nacional",
    "EF": "Educación Física II",
    "FJ": "Reunión FEU/UJC",
    "DI": "Diagnóstico de Inglés",
    "IC": "Acto de Inicio de Curso",
}

subject_types = {
    "PP": "Prueba parcial",
    "Cp": "Control parcial",
    "E": "Evaluación",
    "S": "Seminario",
    "L": "Laboratorio",
    "T": "Taller",
    "&": "Otras actividades",
}

hours = [
    "8:00-9:20",
    "9:30-10:50",
    "11:00-12:20",
    "12:30-1:50",
    "2:00-3:20",
    "4:00-5:20",
]


class SchoolEventType(Enum):
    UNKNOWN = 0,
    CONFERENCE = 1,
    PRACTICAL = 2,
    MIXED = 3,
    OTHER = 4

def parse_event(turn: int, data: str):

    # Si input tiene datos
    if not data: return data
    
    turn_hour = hours[turn-1].split("-")
    start_date = datetime.datetime.strptime(turn_hour[0], "%H:%M")
    end_date = datetime.datetime.strptime(turn_hour[1], "%H:%M")

    start_date = start_date.strftime("%H:%M:%S")
    end_date = end_date.strftime("%H:%M:%S")
    
    data_formatted = {
        "turn": turn,
        "type": "",
        "startDate": start_date,
        "endDate": end_date,
        "name": "",
        "annotations": None,
        "color": None
    }
    
    data_splitted = data.split(",")
    
    # Aquí colocaré las asignaturas ya convertidas
    event_names = []
    event_annotations = []
    
    # Por cada asignatura dividida por comas
    for index in range(0, len(data_splitted)):
        
        # la asignatura actual
        actual_item : str = str(data_splitted[index])

        # si el item tiene un nombre más largo que 4...
        if (len(actual_item) > 4):
            data_formatted['type'] = SchoolEventType.OTHER.name
            event_names.append(actual_item)
            
        else:
            # procedo a convertir usando la leyenda
            
            # Limpio los espacios en blanco
            actual_item = actual_item.replace(" ", "")
            
            # Si el item actual está en las claves de las asignaturas
            if (actual_item.upper() in subject_names.keys()):
                
                event_names.append(subject_names[actual_item.upper()])
                event_annotations.append("Clase")
                
            else:
                
                # de otra forma, es que tiene anotaciones
                
                # Por cada clave en el diccionario de tipos
                for key in subject_types.keys():
                    
                    # Si el item actual comienza con esa clave
                    if actual_item.startswith(key):
                        event_annotations.append(subject_types[key])
                        
                        # Elimina el subject type para que quede solamente
                        # el nombre de la asignatura
                        actual_item = actual_item.replace(key, "")
                        
                        break

                # por cada clave en el diccionario de nombres
                for key in subject_names.keys():
                    
                    # si el item actual termina en la clave
                    if actual_item.upper().endswith(key):
                        
                        event_names.append(subject_names[key])
                        break
            
            # si solo haya una asignatura
            if (len(data_splitted) == 1):
                data_formatted['type'] = SchoolEventType.CONFERENCE.name if (actual_item.upper() == actual_item) else SchoolEventType.PRACTICAL.name
            elif (actual_item.upper() in {"FJ", "DI", "IC"}):
                data_formatted['type'] = SchoolEventType.OTHER.name
            else:
                data_formatted['type'] = SchoolEventType.MIXED.name
                
            
        
    data_formatted.update(
        {
            "name": str.join(" y ", event_names),
            "annotations": str.join(" y ", event_annotations)
        }
    )
    
    return data_formatted


def parse_event2(input: str):

    # Si input tiene datos
    if not input:
        return input

    print(input)

    lsw = input.split(",")
    output = []

    for index in range(0, len(lsw)):
        actual_item = lsw[index]

        actual_item_converted = ["", ""]

        if len(actual_item) > 4:
            actual_item_converted[0] = actual_item
        else:
            actual_item = actual_item.replace(" ", "")

            if (
                actual_item.upper() in subject_names.keys()
                or actual_item.upper() in {"FJ", "DI", "IC"}
            ):
                actual_item_converted[0] = actual_item
            else:
                for key in subject_types.keys():
                    if actual_item.startswith(key):
                        actual_item_converted[1] = key
                        actual_item = actual_item.replace(key, "")
                        break

                for key in subject_names.keys():
                    if actual_item.upper().endswith(key):
                        actual_item_converted[0] = actual_item
                        break

        output.append("/".join(actual_item_converted))

        if output[index] == "/":
            breakpoint()
    return ",".join(output)


def parse_schedule(table_name: str, excel_address: str):
    global wb
    global actual_year

    wb = load_workbook(excel_address)

    title_list = parse_title(str(wb[table_name].cell(1, 1).value))

    actual_year = int(title_list[5].strip())

    output = {
        "scheduleName": "schoolSchedule",
        "author": "Barnés",
        "createdDate": get_actual_day_formatted(),
        "updatedDate": get_actual_day_formatted(),
        "lastTimeDownloaded": "",
        "academy": "Universidad de Holguín",
        "faculty": "FACIM",
        "career": title_list[2].strip(),
        "course": title_list[4].strip(),
        "courseType": title_list[3].strip(),
        "year": title_list[5].strip(),
        "period": title_list[6].strip(),
        "endDate": "",
        "weeks": get_weeks(actual_year, table_name)
    }

    # "scheduleName": title_list[1].strip(),

    # print(get_last_date(output["weeks"]))

    return output


# TODO: Método para extraer las materias directamente de excel


def get_actual_day_formatted():
    return datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def store_schedule_ver2(schedule_name: str, excel_address: str, table_name: str):
    parsed_schedule = parse_schedule(table_name, excel_address)

    # Saving
    with open(f"test.json", "w", encoding="utf8") as fp:
        json.dump(parsed_schedule, fp, ensure_ascii=False, indent=2)


def get_last_date(weeks):
    return weeks[len(weeks) - 1]["weekStart"]


# En Java, el algoritmo utilizado es el Linear Congruential Generator (LCG), mientras que en Python, el algoritmo utilizado es el Mersenne Twister. 
# Estos algoritmos generan números aleatorios de manera diferente, incluso si se les da el mismo seed.
# Además, la forma en que se utiliza el seed en cada lenguaje también puede ser diferente. En Python, el seed se utiliza para inicializar el generador de números aleatorios, mientras que en Java, el seed se utiliza para establecer el estado inicial del generador de números aleatorios. 
# Esto significa que incluso si se utiliza el mismo seed en ambos lenguajes, los generadores de números aleatorios comenzarán en diferentes estados y, por lo tanto, generarán diferentes secuencias de números aleatorios.

def generate_color_from_string(string: str):
    random.seed(string)
    red = random.randint(0, 256)
    green = random.randint(0, 256)
    blue = random.randint(0, 256)
    
    return "#{0:02x}{1:02x}{2:02x}".format(red, green, blue)

# print(generate_color_from_string("hola"))

store_schedule_ver2("facim", "excel/1ro_Ing._Informática Last.xlsx", "1ro")
