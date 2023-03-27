import json
from math import floor
from openpyxl import Workbook, load_workbook

wb = Workbook()

# TODO Hacer un parser para cambiar la forma en la que los turnos son descritos a la forma en la que yo lo uso.


def get_week(week_index, table_name: str) -> dict:

    days = []
    
    
    table = wb[table_name]

    # salto en filas para grupos de 5 semanas
    initial_row = 4 + 8 * floor(week_index / 5)

    # Número de la semana saltando entre cada 5 de estas
    week_number = week_index % 5

    # Posición de inicio de la semana especificada
    week_start_index = 4 + week_number * 5

    for column_index in range(week_start_index, week_start_index+5):

        day = []

        for row_index in range(initial_row, initial_row+6):
            
            
            subject = str(table.cell(row_index, column_index).value) if table.cell(
                row_index, column_index).value != None else ""
            day.append(parse_subject(subject))
            
        days.append(day)
        

    week_start_date = str(table.cell(
        initial_row-2, week_start_index+1).value)
    
    week_start_date_parsed = parse_week_date(
        week_start_date) if week_start_date != "None" else ""

    week_cell_number = str(table.cell(
        initial_row-2, week_start_index).value)
    
    week_cell_number = week_cell_number if week_cell_number != "None" else ""
    result = {
        "number": week_cell_number,
        "weekStart": week_start_date_parsed,
        "days": days
    }

    return result

def parse_week_date(input: str):

    if input != "":
        # Para 20 al 26/feb
        splitted_i = input.strip().split(' ')

        if len(splitted_i[0]) > 2:
            return splitted_i[0]
        else:
            last_item = splitted_i[len(splitted_i) - 1]
            return splitted_i[0] + "/" + last_item.split("/")[1]
    else:
        return input

def get_weeks(table_name):

    is_empty = False
    current_index = 0

    weeks = []
    while is_empty == False:

        current_week = get_week(current_index, table_name)

        if current_week['weekStart'] == "":
            break
        weeks.append(current_week)

        current_index += 1

    return weeks

def parse_title(input: str):
    title = input.strip().split(' ')

    index = 0
    while index < len(title):
        if title[index] == '' and title[index + 1] == '':
            del (title[index])
            index -= 1

        index += 1

    result = ''
    for index in range(len(title)):
        if index > 0:
            result += ' '
        result += parse_name(title[index])

    result_dict = result.split('  ')

    return result_dict

def parse_name(input: str):
    if input.count(':'):
        input = input.split(':')[1].strip()

    return input


subject_names = {
    "M": "Matemática I",
    "P": "Programación I",
    "D": "Matemática Discreta I",
    "I": "Fundamentos de la informática",
    "E": "Electiva",
    "F": "Filosofía",
    "N": "Seguridad Nacional",
    "EF": "Educación Física"
  }

subject_types = {
    "PP": "Prueba parcial",
    "FJ": "Reunión FEU/UJC",
    "DI": "Diagnóstico de Inglés",
    "IC": "Acto de Inicio de Curso",
    "S": "Seminario",
    "L": "Laboratorio",
    "T": "Taller",
    "&": "Otras actividades"
  }

def parse_subject(input : str):
    
    if len(input) > 4: return input
    
    converted_subject = ""
    
    for key in subject_types.keys():
        
        if input.upper().startswith(key):
            
            if key in {"FJ", "DI", "IC"}: return key
            
            converted_subject = key
            input = input.replace(key, "")
            break
    
    for key in subject_names.keys():
        if input.upper().endswith(key):
            
            converted_subject = input + converted_subject
            break
        
    return converted_subject
    
def parse_schedule(table_name: str, excel_address: str):
    
    global wb

    wb = load_workbook(excel_address)

    title_list = parse_title(str(wb[table_name].cell(1, 1).value))

    weeks = get_weeks(table_name)

    output = {
        "scheduleName": title_list[1].strip(),
        "career": title_list[2].strip(),
        "courseType": title_list[3].strip(),
        "year": title_list[4].strip(),
        "course": title_list[5].strip(),
        "period": title_list[6].strip(),
        "updatedDate": "2023-02-23 12:00:00",
        "simbology": {
            "M": "Matemática I",
            "P": "Programación I",
            "D": "Matemática Discreta I",
            "I": "Fundamentos de la informática",
            "E": "Electiva",
            "F": "Filosofía",
            "N": "Seguridad Nacional",
            "EF": "Educación Física"
        },
        "legend": {
            "PP": "Prueba parcial",
            "FJ": "Reunión FEU/UJC",
            "DI": "Diagnóstico de Inglés",
            "IC": "Acto de Inicio de Curso",
            "S": "Seminario",
            "L": "Laboratorio",
            "T": "Taller",
            "&": "Otras actividades"
        },
        "hours": [
            "8:00-9:20",
            "9:30-10:50",
            "11:00-12:20",
            "12:30-1:50",
            "2:00-3:20",
            "4:00-5:20"
        ],

        "weeks": weeks
    }
    return output

# TODO: Método para extraer las materias directamente de excel
def extract_subjects(table_name: str):
    pass

def store_schedule(schedule_name: str, excel_address: str, table_name: str):
    
    parsed_schedule = parse_schedule(table_name, excel_address)
    date = str(parsed_schedule["updatedDate"]).split(" ")[0].replace("-", ".")
    year = str(parsed_schedule["year"]).strip()
    career = str(parsed_schedule["career"]).strip()
    
    # Saving
    with open(f"resources/{schedule_name}/{date}. {year} {career}.json", "w", encoding='utf8') as fp:
        json.dump(parsed_schedule, fp, ensure_ascii=False)


store_schedule("facim", 'excel/1ro I. Informática 23.02.2023.xlsx', "1º")
store_schedule("facim", 'excel/2do I. Informática.xlsx', "2º")

# print(parse_week_date("20/feb al 26/feb"))
# print(parse_week_date("6 al 12/mar"))
# Las casillas completas se cargan desde la primera coincidencia
# print(wb["1º"].cell(12, 5 + 20).value)
