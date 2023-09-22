import datetime
import json
from math import floor
from openpyxl import Workbook, load_workbook

format_version = 1
wb = Workbook()

# TODO Hacer un parser para cambiar la forma en la que los turnos son descritos a la forma en la que yo lo uso.


def get_week(year, week_index, table_name: str) -> dict:
    days = []

    table = wb[table_name]

    # salto en filas para grupos de 5 semanas
    initial_row = 4 + 8 * floor(week_index / 5)

    # Número de la semana saltando entre cada 5 de estas
    week_number = week_index % 5

    # Posición de inicio de la semana especificada
    week_start_index = 4 + week_number * 5

    # Fecha de inicio de semana
    week_start_date = str(table.cell(initial_row - 2, week_start_index + 1).value)
    week_start_date_parsed = (parse_week_date(year, week_start_date) if week_start_date != "None" else "")

    # Número de la semana
    week_cell_number = str(table.cell(initial_row - 2, week_start_index).value)
    week_cell_number = week_cell_number if week_cell_number != "None" else ""

    for column_index in range(week_start_index, week_start_index + 5):
        day = []

        for row_index in range(initial_row, initial_row + 6):
            subject = (
                str(table.cell(row_index, column_index).value)
                if table.cell(row_index, column_index).value != None
                else ""
            )
            day.append(parse_subject(subject))

        days.append(day)

    result = {
        "number": week_cell_number,
        "weekStart": week_start_date_parsed,
        "days": days,
    }

    return result


month_conv = {
    "ene": "1",
    "feb": "2",
    "mar": "3",
    "abr": "4",
    "may": "5",
    "jun": "6",
    "jul": "7",
    "ago": "8",
    "sep": "9",
    "oct": "10",
    "nov": "11",
    "dic": "12",
}

lastMonth = ""
actualYear = 0


def parse_week_date(year: int, input: str):
    global actualYear
    global lastMonth

    if input == "": return input
    
    # Para 20 al 26/feb
    splitted_i = input.strip().split(" ")
    splitted_s = splitted_i[0].split("/")
    splitted_f = splitted_i[2].split("/")

    if len(splitted_s) < 2:
        splitted_s.append(splitted_f[1])

    s_f = [
        splitted_s[0].zfill(2),
        month_conv[splitted_s[1]],
        splitted_f[0].zfill(2),
        month_conv[splitted_f[1]],
    ]

    if lastMonth:
        if lastMonth == "Dec" and s_f[1] == "Jan":
            actualYear = year + 1

    if s_f[2] == "Dec" and s_f[4] == "Jan":
        actualYear = year + 1

    lastMonth = s_f[1]

    s_f.insert(0, str(actualYear))

    final = f"{s_f[0]}-{s_f[2].zfill(2)}-{s_f[1]}"

    print(final)

    return str(final)


def get_weeks(year: int, table_name):
    is_empty = False
    current_index = 0

    weeks = []
    
    while is_empty == False:
        current_week = get_week(year, current_index, table_name)

        if current_week["weekStart"] == "": break
        
        weeks.append(current_week)

        current_index += 1

    return weeks


def parse_title(input: str):
    title = input.strip().split(" ")

    index = 0
    while index < len(title):
        if title[index] == "" and title[index + 1] == "":
            del title[index]
            index -= 1

        index += 1

    result = ""
    for index in range(len(title)):
        if index > 0:
            result += " "
        result += parse_name(title[index])

    result_dict = result.split("  ")

    return result_dict


def parse_name(input: str):
    if input.count(":"):
        input = input.split(":")[1].strip()

    return input


subject_names = {
    "M": "Matemática II",
    "P": "Programación II",
    "D": "Matemática Discreta II",
    "U": "Lenguaje Unificado de Modelado (UML)",
    "H": "Historia de Cuba",
    "N": "Defensa Nacional",
    "EF": "Educación Física II",
}

subject_types = {
    "PP": "Prueba parcial",
    "Cp": "Control parcial",
    "E": "Evaluación",
    "FJ": "Reunión FEU/UJC",
    "DI": "Diagnóstico de Inglés",
    "IC": "Acto de Inicio de Curso",
    "S": "Seminario",
    "L": "Laboratorio",
    "T": "Taller",
    "&": "Otras actividades",
}


def parse_subject(input: str):
    if input:
        lsw = input.split(",")
        output = []

        # print(lsw)
        # if len(input) > 4:
        #     return input

        for index in range(0, len(lsw)):
            actual_item = lsw[index]

            actual_item_converted = ["", ""]

            if len(actual_item) > 4:
                actual_item_converted[0] = actual_item
            else:
                actual_item = actual_item.replace(" ", "")

                if (
                    actual_item.upper() in subject_names.keys()
                    or actual_item.upper() in {"FJ", "DI", "IC"}
                ):
                    actual_item_converted[0] = actual_item
                else:
                    for key in subject_types.keys():
                        if actual_item.startswith(key):
                            actual_item_converted[1] = key
                            actual_item = actual_item.replace(key, "")
                            break

                    for key in subject_names.keys():
                        if actual_item.upper().endswith(key):
                            actual_item_converted[0] = actual_item
                            break

            output.append("/".join(actual_item_converted))

            if output[index] == "/":
                breakpoint()
        return ",".join(output)

    else:
        return input


def parse_schedule(table_name: str, excel_address: str):
    global wb
    global actualYear

    wb = load_workbook(excel_address)

    title_list = parse_title(str(wb[table_name].cell(1, 1).value))

    actualYear = int(title_list[5].strip())
    weeks = get_weeks(actualYear, table_name)

    output = {
        "scheduleName": "schoolSchedule",
        "author": "Barnés",
        "createdDate": get_actual_day_formatted(),
        "updatedDate": get_actual_day_formatted(),
        "lastTimeDownloaded": "",
        "formatVersion": format_version,
        "academy": "Universidad de Holguin",
        "faculty": "FACIM",
        "career": title_list[2].strip(),
        "course": title_list[4].strip(),
        "courseType": title_list[3].strip(),
        "year": title_list[5].strip(),
        "period": title_list[6].strip(),
        "endDate": "",
        "symbology": subject_names,
        "legend": subject_types,
        "hours": [
            "8:00-9:20",
            "9:30-10:50",
            "11:00-12:20",
            "12:30-1:50",
            "2:00-3:20",
            "4:00-5:20",
        ],
        "weeks": weeks
        # "scheduleName": title_list[1].strip(),
    }
    return output


# TODO: Método para extraer las materias directamente de excel


def get_actual_day_formatted():
    return datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def store_schedule_ver1(schedule_name: str, excel_address: str, table_name: str):
    parsed_schedule = parse_schedule(table_name, excel_address)

    # Saving
    with open(f"test.json", "w", encoding="utf8") as fp:
        json.dump(parsed_schedule, fp, ensure_ascii=False, indent=2)


def get_last_date(weeks):
    return weeks[len(weeks) - 1]["weekStart"]

store_schedule_ver1("facim", "excel/1ro 2232.xlsx", "1ro")
